from openpyxl import Workbook, load_workbook
from attendance_handler import get_praise_jam_attendance_array
from openpyxl.styles import Alignment

def init_workbook(name, month, praise_jam_obj):
    wb = Workbook()
    ws = wb.active
    ws.title = name
    #write in for Jam
    cnt_1= 0
    for i in range(3,8):
        curr_obj = praise_jam_obj["T"][cnt_1]
        cnt_1+=1
        ws[F"A{i}"] = curr_obj["FT"]
        ws[F"B{i}"] = curr_obj["ST"]
    #write in for Praise
    cnt_2 = 0
    for i in range(11,16):
        curr_obj = praise_jam_obj["J"][cnt_2]
        ws[F"A{i}"] = curr_obj["FJ"]
        ws[F"B{i}"] = curr_obj["SJ"]
        cnt_2+=1
    cnt_3 = 0
    for i in range(19,24):
        curr_obj = praise_jam_obj["P"][cnt_3]
        ws[F"A{i}"] = curr_obj["FP"]
        ws[F"B{i}"] = curr_obj["SP"]
        cnt_3+=1
    #write in for overall
    for i in range(0,5):
        ws[f"A{i+27}"] = F"=SUM(A{i+3},A{i+11}, A{i+19})"
        ws[f"B{i+27}"] = F"=SUM(B{i+3},B{i+11}, B{i+19})"
    #Insert Averages 
    ws["A1"] = month
    ws["A8"] = "=ROUND(AVERAGE(A3:A7),0)"
    ws["B8"] = "=ROUND(AVERAGE(B3:B7),0)"
    ws["A16"] = "=ROUND(AVERAGE(A11:A15),0)"
    ws["B16"] = "=ROUND(AVERAGE(B11:B15),0)"
    ws["A24"] = "=ROUND(AVERAGE(A19:A23),0)"
    ws["B24"] = "=ROUND(AVERAGE(B19:B23),0)"
    ws["A32"] = "=ROUND(AVERAGE(A27:A31),0)"
    ws["B32"] = "=ROUND(AVERAGE(B27:B31),0)"
    ws["A2"] = "1st"
    ws["A10"] = "1st" 
    ws["A18"] = "1st"
    ws["A26"] = "1st"
    ws["B2"] = "2nd"
    ws["B10"] = "2nd"
    ws["B18"] = "2nd"
    ws["B26"] = "2nd"
    ws.merge_cells('A1:B1')
    ws.merge_cells('A9:B9')
    ws.merge_cells('A17:B17')
    ws.merge_cells('A25:B25')
    #align all center
    for i in range(1,33):
        ws[f"A{i}"].alignment = Alignment(horizontal='center')
        ws[f"B{i}"].alignment = Alignment(horizontal='center')
    
    wb.save(f"{name}.xlsx")

if __name__ == "__main__":
    obj = get_praise_jam_attendance_array("10","2021")
    print(obj)
    init_workbook("test", "Oct", obj)